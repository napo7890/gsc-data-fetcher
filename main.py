#!/usr/bin/python
import pandas as pd
from pandas.io.json import json_normalize
from apiclient.discovery import build
from appengine import config
from google_auth_oauthlib.flow import InstalledAppFlow
import pickle
from openpyxl import load_workbook


def main():
    webmaster = authorize_access()
    read_site_list(webmaster)

    for site, site_url in config.site_list.items():
        gsc_data_date, gsc_data_page_date, gsc_data_query, gsc_data_device, gsc_data_page_query = query_data(webmaster, site_url)
        write_file(site, '''site_list''', gsc_data_date, gsc_data_page_date, gsc_data_query, gsc_data_device,
                   gsc_data_page_query)
    return site, site_url


def authorize_access():
    try:
        credentials = pickle.load(open("appengine/credentials.pickle", "rb"))
    except (OSError, IOError) as e:
        flow = InstalledAppFlow.from_client_secrets_file(config.CLIENT_SECRET_JSON, scopes=config.OAUTH_SCOPE)
        credentials = flow.run_console()
        pickle.dump(credentials, open("appengine/credentials.pickle", "wb"))

    webmasters_service = build('webmasters', 'v3', credentials=credentials)
    return webmasters_service


def read_site_list(webmasters_service):
    # Retrieve list of properties in account
    site_list = webmasters_service.sites().list().execute()

    # Filter for verified websites
    verified_sites_urls = [s['siteUrl'] for s in site_list['siteEntry']
                           if s['permissionLevel'] != 'siteUnverifiedUser'
                           and s['siteUrl'][:4] == 'http']

    # Printing & saving the URLs of all websites you are verified for.
    site_list = []
    for site_url in verified_sites_urls:
        site_list.append(site_url)
    return site_list


def query_data(webmasters_service, site_url):
    # Query all data
    gsc_data_date = webmasters_service.searchanalytics().query(siteUrl=site_url, body=config.GSC_QUERY_DATE).execute()
    gsc_data_date = json_normalize(gsc_data_date['rows'])
    df_keys__date = gsc_data_date['keys'].apply(pd.Series).fillna(0)
    gsc_data_date = pd.concat([gsc_data_date, df_keys__date], axis=1)

    gsc_data_page_date = webmasters_service.searchanalytics().query(siteUrl=site_url, body=config.GSC_QUERY_PAGE_DATE).execute()
    gsc_data_page_date = json_normalize(gsc_data_page_date['rows'])
    df_keys_page_date = gsc_data_page_date['keys'].apply(pd.Series).fillna(0)
    gsc_data_page_date = pd.concat([gsc_data_page_date, df_keys_page_date], axis=1)

    gsc_data_query_date = webmasters_service.searchanalytics().query(siteUrl=site_url, body=config.GSC_QUERY_QUERY_DATE).execute()
    gsc_data_query_date = json_normalize(gsc_data_query_date['rows'])
    df_keys_query_date = gsc_data_query_date['keys'].apply(pd.Series).fillna(0)
    gsc_data_query_date = pd.concat([gsc_data_query_date, df_keys_query_date], axis=1)

    gsc_data_device = webmasters_service.searchanalytics().query(siteUrl=site_url, body=config.GSC_QUERY_DEVICE).execute()
    gsc_data_device = json_normalize(gsc_data_device['rows'])
    df_keys_gsc_data_device = gsc_data_device['keys'].apply(pd.Series).fillna(0)
    gsc_data_device = pd.concat([gsc_data_device, df_keys_gsc_data_device], axis=1)

    gsc_data_page_query_date = webmasters_service.searchanalytics().query(siteUrl=site_url, body=config.GSC_QUERY_QUERY_DATE_PAGE).execute()
    gsc_data_page_query_date = json_normalize(gsc_data_page_query_date['rows'])
    df_keys_page_query_date = gsc_data_page_query_date['keys'].apply(pd.Series).fillna(0)
    gsc_data_page_query_date = pd.concat([gsc_data_page_query_date, df_keys_page_query_date], axis=1)

    return gsc_data_date, gsc_data_page_date, gsc_data_query_date, gsc_data_device, gsc_data_page_query_date


def write_file(site, gsc_data_date, gsc_data_page_date, gsc_data_query_date, gsc_data_device, gsc_data_page_query_date):
    filename = 'data/' + site + '-' + 'gsc-data-backup.xlsx'

    df_gsc_data_date = pd.DataFrame(gsc_data_date).set_index('keys')
    df_gsc_data_date.columns = ['Clicks', 'CTR', 'Impressions', 'Position', 'Date']

    df_gsc_data_page_date = pd.DataFrame(gsc_data_page_date).set_index('keys')
    df_gsc_data_page_date.columns = ['Clicks', 'CTR', 'Impressions', 'Position', 'Page', 'Date']

    df_gsc_data_query_date = pd.DataFrame(gsc_data_query_date).set_index('keys')
    df_gsc_data_query_date.columns = ['Clicks', 'CTR', 'Impressions', 'Position', 'Query', 'Date']

    df_gsc_data_device = pd.DataFrame(gsc_data_device).set_index('keys')
    df_gsc_data_device.columns = ['Clicks', 'CTR', 'Impressions', 'Position', 'Platform']

    df_gsc_data_page_query_date = pd.DataFrame(gsc_data_page_query_date).set_index('keys')
    df_gsc_data_page_query_date.columns = ['Clicks', 'CTR', 'Impressions', 'Position', 'Query', 'Date', 'Page']

    # dropping null value columns to avoid errors
    gsc_data_page_query_date.dropna(inplace=True)

    append_df_to_excel(filename, df_gsc_data_date, sheet_name='Data by Date', index=False, header=False)
    append_df_to_excel(filename, df_gsc_data_page_date, sheet_name='Data by Page Date', index=False, header=False)
    append_df_to_excel(filename, df_gsc_data_query_date, sheet_name='Data by Query Date', index=False, header=False)
    append_df_to_excel(filename, df_gsc_data_page_query_date, sheet_name='Data by Query Date Page', index=False, header=False)
    # append_df_to_excel(filename, df_gsc_data_device, sheet_name='Data by Device', index=False, header=False)


def append_df_to_excel(filename, df, sheet_name='sheet_name', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        '''
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        '''
        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


if __name__ == '__main__': main()
